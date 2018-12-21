# Example
'''
Modify here
'''
libraryExcelPath = r'C:\Users\khanhdangnguyen\Documents\ExcelPackage'
workingPath = r'C:\Users\khanhdangnguyen\Documents\03_Document'
versionColumn = 'B'
startRow = 3
sheetCheck = 'History'
import numbers
# import PIL

def is_number(s):
	try:
		float(s)
		return True
	except ValueError:
		return False

def checkInFolder(folder):
	os.chdir('.\\'+folder)
	logFile.write(os.getcwd()+'\n')
	fileList = os.listdir(".")
	for item in fileList:
		print(item)
		if os.path.isdir('.\\'+item):
			checkInFolder(item)
			os.chdir('..')
			logFile.write(os.getcwd()+'\n')
		elif item.endswith('.xlsx'):
			excelItem = item.ljust(40)
			logFile.write(excelItem)
			wb = openpyxl.load_workbook(item)
			if sheetCheck not in wb.sheetnames:
				logFile.write('Err'.rjust(5)+'\n')
			else:
				sheet = wb.get_sheet_by_name(sheetCheck)
				i = startRow
				version = 'None'
				for row in range(startRow,sheet.max_row + 1):
					if is_number(str(sheet[versionColumn+str(row)].value)):
						version = str(sheet[versionColumn+str(row)].value)
				logFile.write(version.rjust(5)+'\n')
		elif item.endswith('.xls'):
			excelItem = item.ljust(40)
			logFile.write(excelItem)
			ob = xlrd.open_workbook(item)
			if sheetCheck in ob.sheet_names():
				osheet = ob.sheet_by_index(ob.sheet_names().index(sheetCheck))
				# maxRow = max(osheet.col_values(1))
				print(osheet.nrows)
				for row in range(startRow,osheet.nrows):
					if is_number(osheet.cell(row,1).value):
						version = str(osheet.cell(row,1).value)
				logFile.write(version.rjust(5)+'\n')
			else:
				logFile.write('Err'.rjust(5)+'\n')
	# os.chdir('..')


import os
currentPath = os.getcwd()
print('Current working directory: '+os.getcwd())
print('Move to library directory: '+libraryExcelPath)
os.chdir(libraryExcelPath)
# os.chdir(r'.\et_xmlfile-1.0.1')
import et_xmlfile
# os.chdir(r'..\jdcal-1.4')
import jdcal
# os.chdir(r'..\openpyxl')
import openpyxl
os.chdir(r'.\xlrd')
import xlrd
# os.chdir(r'..\Pillow')
# import Pillow
os.chdir(r'..')

os.chdir(workingPath)
print('Move to directory: '+workingPath)

fileList = os.listdir(".")
import pprint
#pprint.pprint(fileList)

logFile = open(r'.\log.txt','w')
logFile.write(os.getcwd()+'\n')
logFile.close()

logFile = open(r'.\log.txt','a')
checkInFolder('.')			
logFile.close()
	
os.chdir(workingPath)
print('Current working directory: '+workingPath)
logFile = open(r'.\log.txt')
print(logFile.read())
logFile.close()

print('Move back to library directory: '+currentPath)
os.chdir(currentPath)

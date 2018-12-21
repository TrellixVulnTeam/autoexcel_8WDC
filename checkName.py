# Example
'''
Modify here
'''
libraryExcelPath = r'C:\Users\khanhdangnguyen\Documents\ExcelPackage'
workingPath = r'C:\Users\khanhdangnguyen\Documents\03_Document'
versionColumn = 'B'
authorColumn = 'F'
checkerColumn = 'E'
approverColumn = 'D'
startRow = 3
sheetCheck = 'Version'
import numbers
# import PIL

import re
dateRegex = re.compile(r'\d\d/\d\d/\d\d\d\d')
def is_number(s):
	try:
		float(s)
		return True
	except ValueError:
		return False

def checkInFolder(folder):
	os.chdir('.\\'+folder)
	logFile.write(os.getcwd()+'\n')
	fileList = os.listdir(".")
	for item in fileList:
		print(item)
		if os.path.isdir('.\\'+item):
			checkInFolder(item)
			os.chdir('..')
			logFile.write(os.getcwd()+'\n')
		elif item.endswith('.xlsx'):
			excelItem = item.ljust(40)
			logFile.write(excelItem)
			wb = openpyxl.load_workbook(item)
			if sheetCheck not in wb.sheetnames:
				logFile.write('Err'.rjust(5)+'\n')
			else:
				sheet = wb.get_sheet_by_name(sheetCheck)
				i = startRow
				version = 'None'
				for row in range(startRow,sheet.max_row + 1):
					if is_number(str(sheet[versionColumn+str(row)].value)):
						version = str(sheet[versionColumn+str(row)].value)
						approver = str(sheet[approverColumn+str(row)].value).partition("\n")[0]
						checker = str(sheet[checkerColumn+str(row)].value).partition("\n")[0]
						author = str(sheet[authorColumn+str(row)].value).partition("\n")[0]
				logFile.write(version.rjust(5))
				logFile.write(author.rjust(15))
				logFile.write(checker.rjust(15))
				logFile.write(approver.rjust(15))
				for row in range(startRow,sheet.max_row + 1):
					if dateRegex.search(str(sheet[approverColumn+str(row)].value)) is not None:
						date = dateRegex.search(str(sheet[approverColumn+str(row)].value))
						logFile.write(date.group().rjust(15))
				
				logFile.write('\n')
				
		elif item.endswith('.xls'):
			excelItem = item.ljust(40)
			logFile.write(excelItem)
			ob = xlrd.open_workbook(item)
			if sheetCheck in ob.sheet_names():
				osheet = ob.sheet_by_index(ob.sheet_names().index(sheetCheck))
				# maxRow = max(osheet.col_values(1))
				print(osheet.nrows)
				for row in range(startRow-1,osheet.nrows):
					if is_number(osheet.cell(row,1).value):
						version = str(osheet.cell(row,1).value)
				logFile.write(version.rjust(5))
				
				for row in range(startRow-1,osheet.nrows):
					if is_number(osheet.cell(row,1).value):
						version = str(osheet.cell(row,1).value)
						approver = str(osheet.cell(row,3).value).partition("\n")[0]
						checker = str(osheet.cell(row,4).value).partition("\n")[0]
						author = str(osheet.cell(row,5).value).partition("\n")[0]
				# logFile.write(version.rjust(5))
				logFile.write(author.rjust(15))
				logFile.write(checker.rjust(15))
				logFile.write(approver.rjust(15))
				for row in range(startRow-1,osheet.nrows):
					if dateRegex.search(str(osheet.cell(row,3).value)) is not None:
						date = dateRegex.search(str(osheet.cell(row,3).value))
						logFile.write(date.group().rjust(15))
				
				logFile.write('\n')
			else:
				logFile.write('Err'.rjust(5)+'\n')
	# os.chdir('..')


import os
currentPath = os.getcwd()
print('Current working directory: '+os.getcwd())
print('Move to library directory: '+libraryExcelPath)
os.chdir(libraryExcelPath)
# os.chdir(r'.\et_xmlfile-1.0.1')
import et_xmlfile
# os.chdir(r'..\jdcal-1.4')
import jdcal
# os.chdir(r'..\openpyxl')
import openpyxl
os.chdir(r'.\xlrd')
import xlrd
# os.chdir(r'..\Pillow')
# import Pillow
os.chdir(r'..')

os.chdir(workingPath)
print('Move to directory: '+workingPath)

fileList = os.listdir(".")
import pprint
#pprint.pprint(fileList)

logFile = open(r'.\log.txt','w')
logFile.write(os.getcwd()+'\n')
logFile.close()

logFile = open(r'.\log.txt','a')
checkInFolder('.')			
logFile.close()
	
os.chdir(workingPath)
print('Current working directory: '+workingPath)
logFile = open(r'.\log.txt')
print(logFile.read())
logFile.close()

print('Move back to library directory: '+currentPath)
os.chdir(currentPath)

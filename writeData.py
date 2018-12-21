libraryExcelPath = r'C:\Users\khanhdangnguyen\Documents\ExcelPackage'
workingPath = r'C:\Users\khanhdangnguyen\Documents\03_Document'
versionColumn = 'B'
authorColumn = 'F'
checkerColumn = 'E'
approverColumn = 'D'
startRow = 3
sheetCheck = 'Version'
import numbers
# import PIL
import os
currentPath = os.getcwd()
print('Current working directory: '+os.getcwd())
print('Move to library directory: '+libraryExcelPath)
os.chdir(libraryExcelPath)
# os.chdir(r'.\et_xmlfile-1.0.1')
import et_xmlfile
# os.chdir(r'..\jdcal-1.4')
import jdcal
# os.chdir(r'..\openpyxl')
import openpyxl
os.chdir(r'.\xlrd')
# import xlrd
from xlutils.copy import copy
from xlrd import open_workbook
# os.chdir(r'..\Pillow')
# import Pillow
os.chdir(r'..')

os.chdir(workingPath)
print('Move to directory: '+workingPath)

fileList = os.listdir(".")
import pprint
#pprint.pprint(fileList)
inputFile = open(r'.\input.txt')
outputFile = open(r'.\output.txt','w')
listFile = open(r'.\list.txt','w')
outputFile.write(os.getcwd()+'\n')
outputFile.close()
outputFile = open(r'.\output.txt','a')
dictExcel={}
line = inputFile.readline()
cnt = 1
while line:
	listLine = line.strip().split()
	outputFile.write(listLine[0].ljust(25))
	outputFile.write(listLine[1].ljust(25))
	print(listLine[0]+'      '+listLine[1])
	if len(listLine) > 2:
		outputFile.write('Have space in file!!'.rjust(25))
	elif listLine[1] not in fileList:
		outputFile.write('Do not have that file'.rjust(25))
	elif listLine[1].endswith('.xlsx'):
		# wb = openpyxl.load_workbook(listLine[1])
		
		# for sheetCheck in ['Revision','History','Version']:
			# if sheetCheck in wb.sheetnames:
				# sheet = wb.get_sheet_by_name(sheetCheck)
				# sheet['H2'].value = listLine[0]
				# outputFile.write('Done'.rjust(25))
				# wb.save(listLine[1])
		dictExcel[listLine[1]] = listLine[0]
		# outputFile.write('Done'.rjust(25))
		
	elif listLine[1].endswith('.xls'):
		outputFile.write('Not yet')
		listFile.write(listLine[1]+'\n')
		dictExcel[listLine[1]] = listLine[0]
	
	outputFile.write('\n')
	line = inputFile.readline()

for file in dictExcel.keys():
	if file in fileList and file.endswith('.xlsx'):
		outputFile.write('File name '+file)
		print('File name '+file)
		try:
			wb = openpyxl.load_workbook(file)
			outputFile.write('    Okay ')
			print('    Okay ')
			for sheetCheck in ['Revision','History','Version']:
				if sheetCheck in wb.sheetnames:
					sheet = wb.get_sheet_by_name(sheetCheck)
					sheet['H2'].value = dictExcel[file]
					outputFile.write('Done'.rjust(25))
			wb.save(file)
		except KeyError:
			outputFile.write('     Error ')
			print('     Error ')
			
		outputFile.write('\n')
	
	# if file in fileList and file.endswith('.xls'):
		# outputFile.write('File name '+file)
		# print('File name '+file)
		# try:
			# rb = open_workbook(file)
			# wb = copy(rb)
			# outputFile.write('    Okay ')
			# print('    Okay ')
			# for sheetCheck in ['Revision','History','Version']:
				# if sheetCheck in wb.sheetnames:
					# sheet = wb.get_sheet_by_name(sheetCheck)
					# sheet['H2'].value = dictExcel[file]
					# outputFile.write('Done'.rjust(25))
			# wb.save(file)
			# s = wb.get_sheet(0)
			# s.write(7,8,dictExcel[file])
			# wb.save('.\\output\\'+file)
		# except KeyError:
			# outputFile.write('     Error ')
			# print('     Error ')
			
		# outputFile.write('\n')
		
listFile.close()
inputFile.close()
outputFile.close()


		
	
os.chdir(workingPath)
print('Current working directory: '+workingPath)
outputFile = open(r'.\log.txt')
print(outputFile.read())
outputFile.close()

print('Move back to library directory: '+currentPath)
os.chdir(currentPath)


